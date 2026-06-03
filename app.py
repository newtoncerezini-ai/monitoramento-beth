from __future__ import annotations

import argparse
import csv
import re
import os
import shutil
import sqlite3
from datetime import date, datetime
from functools import wraps
from pathlib import Path

from flask import Flask, Response, flash, g, redirect, render_template, request, session, url_for
from openpyxl import load_workbook
from werkzeug.security import check_password_hash, generate_password_hash


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
LOCAL_DB_PATH = DATA_DIR / "monitoramento.db"
DB_PATH = Path("/tmp/monitoramento.db") if os.environ.get("VERCEL") else LOCAL_DB_PATH
DEFAULT_IMPORT = Path(r"C:\Users\newton.cerezini\Downloads\COMFEM - BD PE26271.xlsx")

STATUS_OPTIONS = ["Nao iniciado", "Em andamento", "Concluido", "Atrasado", "Pausado", "Cancelado"]
ROLE_OPTIONS = ["admin", "gestor", "usuario"]
WRITE_ROLES = {"admin", "gestor"}

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "monitoramento-local-dev")


def prepare_db_file() -> None:
    if os.environ.get("VERCEL"):
        if not DB_PATH.exists() and LOCAL_DB_PATH.exists():
            shutil.copyfile(LOCAL_DB_PATH, DB_PATH)
        return
    DATA_DIR.mkdir(exist_ok=True)


def get_db() -> sqlite3.Connection:
    if "db" not in g:
        prepare_db_file()
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        g.db = conn
    return g.db


@app.teardown_appcontext
def close_db(_error: Exception | None) -> None:
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db() -> None:
    prepare_db_file()
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS people (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            email TEXT,
            area TEXT,
            active INTEGER NOT NULL DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS actions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT NOT NULL UNIQUE,
            parent_id INTEGER REFERENCES actions(id) ON DELETE SET NULL,
            level INTEGER NOT NULL CHECK(level BETWEEN 1 AND 3),
            title TEXT NOT NULL,
            area TEXT,
            why TEXT,
            planned_start DATE,
            planned_end DATE,
            actual_start DATE,
            actual_end DATE,
            place TEXT,
            method TEXT,
            cost TEXT,
            status TEXT NOT NULL DEFAULT 'Nao iniciado',
            observations TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS action_people (
            action_id INTEGER NOT NULL REFERENCES actions(id) ON DELETE CASCADE,
            person_id INTEGER NOT NULL REFERENCES people(id) ON DELETE CASCADE,
            role TEXT NOT NULL DEFAULT 'Responsavel',
            PRIMARY KEY (action_id, person_id)
        );

        CREATE TABLE IF NOT EXISTS updates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            action_id INTEGER NOT NULL REFERENCES actions(id) ON DELETE CASCADE,
            update_date DATE NOT NULL,
            status TEXT NOT NULL,
            progress INTEGER NOT NULL DEFAULT 0 CHECK(progress BETWEEN 0 AND 100),
            comment TEXT,
            next_step TEXT,
            blocker TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );

        CREATE INDEX IF NOT EXISTS idx_actions_parent ON actions(parent_id);
        CREATE INDEX IF NOT EXISTS idx_actions_status ON actions(status);
        CREATE INDEX IF NOT EXISTS idx_actions_planned_end ON actions(planned_end);

        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'usuario',
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
        """
    )
    admin = conn.execute("SELECT id FROM users WHERE username = 'admin'").fetchone()
    if admin is None:
        conn.execute(
            "INSERT INTO users (name, username, password_hash, role) VALUES (?, ?, ?, ?)",
            ("Administrador", "admin", generate_password_hash("admin123"), "admin"),
        )
    conn.commit()
    conn.close()


def db() -> sqlite3.Connection:
    init_db()
    return get_db()


@app.before_request
def load_current_user() -> None:
    init_db()
    g.current_user = None
    user_id = session.get("user_id")
    if user_id:
        g.current_user = get_db().execute(
            "SELECT * FROM users WHERE id = ? AND active = 1",
            (user_id,),
        ).fetchone()


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if g.current_user is None:
            return redirect(url_for("login", next=request.path))
        return view(*args, **kwargs)

    return wrapped


def admin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if g.current_user is None:
            return redirect(url_for("login", next=request.path))
        if g.current_user["role"] != "admin":
            flash("Acesso restrito a administradores.")
            return redirect(url_for("dashboard"))
        return view(*args, **kwargs)

    return wrapped


def write_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if g.current_user is None:
            return redirect(url_for("login", next=request.path))
        if g.current_user["role"] not in WRITE_ROLES:
            flash("Seu perfil permite consulta, mas nao permite cadastrar ou alterar dados.")
            return redirect(url_for("dashboard"))
        return view(*args, **kwargs)

    return wrapped


@app.context_processor
def inject_permissions():
    user = getattr(g, "current_user", None)
    can_write = bool(user and user["role"] in WRITE_ROLES)
    can_admin = bool(user and user["role"] == "admin")
    return {"can_write": can_write, "can_admin": can_admin}


def parse_date(value) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def normalize_code(value) -> str:
    text = str(value or "").strip().replace(",", ".")
    text = re.sub(r"\.+$", "", text)
    text = re.sub(r"\s+", "", text)
    return text


def action_level(code: str) -> int:
    return min(code.count(".") + 1, 3)


def parent_code(code: str) -> str | None:
    parts = code.split(".")
    if len(parts) <= 1:
        return None
    return ".".join(parts[:-1])


def split_people(text: str | None) -> list[str]:
    if not text:
        return []
    value = str(text).replace("\n", "/")
    value = re.sub(r"\s+e\s+", "/", value, flags=re.IGNORECASE)
    pieces = re.split(r"[/;,]", value)
    names = []
    for piece in pieces:
        name = re.sub(r"\s+", " ", piece).strip()
        if name and name not in names:
            names.append(name)
    return names


def status_for_row(raw_status: str | None, planned_end: str | None, actual_end: str | None) -> str:
    if raw_status and str(raw_status).strip():
        return str(raw_status).strip()
    if actual_end:
        return "Concluido"
    if planned_end and planned_end < date.today().isoformat():
        return "Atrasado"
    return "Nao iniciado"


@app.route("/login", methods=["GET", "POST"])
def login():
    if g.current_user is not None:
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = db().execute(
            "SELECT * FROM users WHERE username = ? AND active = 1",
            (username,),
        ).fetchone()
        if user and check_password_hash(user["password_hash"], password):
            session.clear()
            session["user_id"] = user["id"]
            return redirect(request.args.get("next") or url_for("dashboard"))
        flash("Usuario ou senha invalidos.")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/admin/users", methods=["GET", "POST"])
@admin_required
def admin_users():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        role = request.form.get("role") if request.form.get("role") in ROLE_OPTIONS else "usuario"
        if not name or not username or not password:
            flash("Preencha nome, usuario e senha.")
            return redirect(url_for("admin_users"))
        try:
            db().execute(
                "INSERT INTO users (name, username, password_hash, role) VALUES (?, ?, ?, ?)",
                (name, username, generate_password_hash(password), role),
            )
            db().commit()
            flash("Usuario criado com sucesso.")
        except sqlite3.IntegrityError:
            flash("Ja existe um usuario com esse login.")
        return redirect(url_for("admin_users"))

    users = db().execute("SELECT * FROM users ORDER BY active DESC, name").fetchall()
    return render_template("admin_users.html", users=users, role_options=ROLE_OPTIONS)


@app.route("/admin")
@write_required
def admin_home():
    actions_count = db().execute("SELECT COUNT(*) AS total FROM actions").fetchone()["total"]
    people_count = db().execute("SELECT COUNT(*) AS total FROM people WHERE active = 1").fetchone()["total"]
    users_count = db().execute("SELECT COUNT(*) AS total FROM users WHERE active = 1").fetchone()["total"]
    return render_template(
        "admin_home.html",
        actions_count=actions_count,
        people_count=people_count,
        users_count=users_count,
    )


@app.route("/admin/users/<int:user_id>/toggle", methods=["POST"])
@admin_required
def admin_user_toggle(user_id: int):
    if user_id == g.current_user["id"]:
        flash("Voce nao pode desativar o proprio usuario.")
        return redirect(url_for("admin_users"))
    db().execute("UPDATE users SET active = CASE active WHEN 1 THEN 0 ELSE 1 END WHERE id = ?", (user_id,))
    db().commit()
    flash("Situacao do usuario atualizada.")
    return redirect(url_for("admin_users"))


@app.route("/admin/users/<int:user_id>/password", methods=["POST"])
@admin_required
def admin_user_password(user_id: int):
    password = request.form.get("password", "")
    if len(password) < 6:
        flash("A senha deve ter pelo menos 6 caracteres.")
        return redirect(url_for("admin_users"))
    db().execute("UPDATE users SET password_hash = ? WHERE id = ?", (generate_password_hash(password), user_id))
    db().commit()
    flash("Senha atualizada.")
    return redirect(url_for("admin_users"))


def upsert_person(conn: sqlite3.Connection, name: str, area: str | None = None) -> int:
    conn.execute(
        "INSERT OR IGNORE INTO people (name, area) VALUES (?, ?)",
        (name, area),
    )
    if area:
        conn.execute("UPDATE people SET area = COALESCE(NULLIF(area, ''), ?) WHERE name = ?", (area, name))
    row = conn.execute("SELECT id FROM people WHERE name = ?", (name,)).fetchone()
    return int(row["id"])


def import_excel(path: Path) -> tuple[int, int]:
    init_db()
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    wb = load_workbook(path, data_only=True)
    ws = wb["TODOS"] if "TODOS" in wb.sheetnames else wb.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    imported = 0
    people_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        values = dict(zip(headers, row))
        if not any(v not in (None, "") for v in row):
            continue
        code = normalize_code(values.get("#"))
        title = values.get("O que? (what)")
        if not code or not title:
            continue

        planned_start = parse_date(values.get("Início Planejado"))
        planned_end = parse_date(values.get("Fim Planejado"))
        actual_start = parse_date(values.get("Início Real"))
        actual_end = parse_date(values.get("Fim Real"))
        status = status_for_row(values.get("Status"), planned_end, actual_end)
        if status not in STATUS_OPTIONS:
            status = "Nao iniciado"

        parent = parent_code(code)
        parent_id = None
        if parent:
            parent_row = conn.execute("SELECT id FROM actions WHERE code = ?", (parent,)).fetchone()
            parent_id = parent_row["id"] if parent_row else None

        conn.execute(
            """
            INSERT INTO actions (
                code, parent_id, level, title, area, why, planned_start, planned_end,
                actual_start, actual_end, place, method, cost, status, observations, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(code) DO UPDATE SET
                parent_id=excluded.parent_id,
                level=excluded.level,
                title=excluded.title,
                area=excluded.area,
                why=excluded.why,
                planned_start=excluded.planned_start,
                planned_end=excluded.planned_end,
                actual_start=excluded.actual_start,
                actual_end=excluded.actual_end,
                place=excluded.place,
                method=excluded.method,
                cost=excluded.cost,
                status=excluded.status,
                observations=excluded.observations,
                updated_at=CURRENT_TIMESTAMP
            """,
            (
                code,
                parent_id,
                action_level(code),
                str(title).strip(),
                values.get("Área"),
                values.get("Por que? (why)"),
                planned_start,
                planned_end,
                actual_start,
                actual_end,
                values.get("Onde? (where)"),
                values.get("Como? (how)"),
                values.get("Quanto? (how much)"),
                status,
                values.get("Observações"),
            ),
        )
        action_id = conn.execute("SELECT id FROM actions WHERE code = ?", (code,)).fetchone()["id"]
        conn.execute("DELETE FROM action_people WHERE action_id = ?", (action_id,))
        for name in split_people(values.get("Nome")):
            person_id = upsert_person(conn, name, values.get("Área"))
            conn.execute(
                "INSERT OR IGNORE INTO action_people (action_id, person_id) VALUES (?, ?)",
                (action_id, person_id),
            )
            people_count += 1
        imported += 1

    conn.commit()
    conn.close()
    return imported, people_count


def rows_to_actions(rows: list[sqlite3.Row]) -> list[dict]:
    actions = []
    for row in rows:
        item = dict(row)
        item["people"] = row["people"].split(" | ") if row["people"] else []
        actions.append(item)
    return actions


def query_actions(filters: dict | None = None) -> list[dict]:
    filters = filters or {}
    clauses = []
    params = []
    if filters.get("status"):
        clauses.append("a.status = ?")
        params.append(filters["status"])
    if filters.get("person_id"):
        clauses.append("a.id IN (SELECT action_id FROM action_people WHERE person_id = ?)")
        params.append(filters["person_id"])
    if filters.get("q"):
        clauses.append("(a.title LIKE ? OR a.code LIKE ? OR a.observations LIKE ?)")
        like = f"%{filters['q']}%"
        params.extend([like, like, like])
    where = "WHERE " + " AND ".join(clauses) if clauses else ""
    rows = db().execute(
        f"""
        SELECT a.*, p.title AS parent_title, GROUP_CONCAT(pe.name, ' | ') AS people
        FROM actions a
        LEFT JOIN actions p ON p.id = a.parent_id
        LEFT JOIN action_people ap ON ap.action_id = a.id
        LEFT JOIN people pe ON pe.id = ap.person_id
        {where}
        GROUP BY a.id
        ORDER BY
            CAST(substr(a.code, 1, instr(a.code || '.', '.') - 1) AS INTEGER),
            a.code
        """,
        params,
    ).fetchall()
    return rows_to_actions(rows)


def get_people() -> list[sqlite3.Row]:
    return db().execute("SELECT * FROM people WHERE active = 1 ORDER BY name").fetchall()


def action_is_overdue(action: dict, today: str | None = None) -> bool:
    today = today or date.today().isoformat()
    return bool(action["planned_end"] and action["planned_end"] < today and not action["actual_end"] and action["status"] != "Concluido")


def action_due_soon(action: dict, days: int = 7) -> bool:
    if not action["planned_end"] or action["actual_end"] or action["status"] == "Concluido":
        return False
    try:
        due = datetime.strptime(action["planned_end"], "%Y-%m-%d").date()
    except ValueError:
        return False
    delta = (due - date.today()).days
    return 0 <= delta <= days


def completion_rate(actions: list[dict]) -> int:
    if not actions:
        return 0
    done = len([a for a in actions if a["status"] == "Concluido"])
    return round((done / len(actions)) * 100)


@app.route("/")
@login_required
def dashboard():
    actions = query_actions()
    total = len(actions)
    overdue = [a for a in actions if a["status"] == "Atrasado" or (a["planned_end"] and a["planned_end"] < date.today().isoformat() and not a["actual_end"])]
    done = [a for a in actions if a["status"] == "Concluido"]
    planned = [a for a in actions if a["planned_end"]]
    next_actions = sorted([a for a in planned if a["status"] != "Concluido"], key=lambda a: a["planned_end"])[:8]
    by_status = {status: len([a for a in actions if a["status"] == status]) for status in STATUS_OPTIONS}
    plans = [a for a in actions if a["level"] == 1]
    return render_template(
        "dashboard.html",
        total=total,
        overdue=len(overdue),
        done=len(done),
        planned=len(planned),
        next_actions=next_actions,
        by_status=by_status,
        plans=plans,
        today=date.today().isoformat(),
    )


@app.route("/actions")
@login_required
def actions_index():
    filters = {
        "status": request.args.get("status", ""),
        "person_id": request.args.get("person_id", ""),
        "q": request.args.get("q", ""),
    }
    return render_template(
        "actions.html",
        actions=query_actions(filters),
        people=get_people(),
        status_options=STATUS_OPTIONS,
        filters=filters,
        today=date.today().isoformat(),
    )


@app.route("/monitoring/alerts")
@login_required
def monitoring_alerts():
    actions = query_actions()
    overdue = [a for a in actions if action_is_overdue(a)]
    due_soon = [a for a in actions if action_due_soon(a)]
    without_owner = [a for a in actions if not a["people"]]
    return render_template(
        "monitoring_alerts.html",
        overdue=overdue,
        due_soon=due_soon,
        without_owner=without_owner,
        today=date.today().isoformat(),
    )


@app.route("/monitoring/calendar")
@login_required
def monitoring_calendar():
    actions = [a for a in query_actions() if a["planned_end"]]
    actions = sorted(actions, key=lambda a: a["planned_end"])
    months = {}
    for action in actions:
        month_key = action["planned_end"][:7]
        months.setdefault(month_key, []).append(action)
    return render_template("monitoring_calendar.html", months=months, today=date.today().isoformat())


@app.route("/monitoring/people")
@login_required
def monitoring_people():
    rows = db().execute(
        """
        SELECT
            p.id,
            p.name,
            p.area,
            COUNT(a.id) AS total,
            SUM(CASE WHEN a.status = 'Concluido' THEN 1 ELSE 0 END) AS completed,
            SUM(CASE WHEN a.status = 'Em andamento' THEN 1 ELSE 0 END) AS in_progress,
            SUM(CASE WHEN (a.status = 'Atrasado' OR (a.planned_end < DATE('now') AND a.actual_end IS NULL AND a.status != 'Concluido')) THEN 1 ELSE 0 END) AS overdue
        FROM people p
        LEFT JOIN action_people ap ON ap.person_id = p.id
        LEFT JOIN actions a ON a.id = ap.action_id
        WHERE p.active = 1
        GROUP BY p.id
        ORDER BY overdue DESC, total DESC, p.name
        """
    ).fetchall()
    return render_template("monitoring_people.html", rows=rows)


@app.route("/monitoring/plans")
@login_required
def monitoring_plans():
    actions = query_actions()
    plans = []
    for plan in [a for a in actions if a["level"] == 1]:
        related = [a for a in actions if a["code"] == plan["code"] or a["code"].startswith(plan["code"] + ".")]
        overdue = [a for a in related if action_is_overdue(a)]
        critical = sorted([a for a in related if a["level"] > 1 and (action_is_overdue(a) or action_due_soon(a))], key=lambda a: a["planned_end"] or "9999-12-31")[:5]
        plans.append(
            {
                "plan": plan,
                "total": len(related),
                "completed": len([a for a in related if a["status"] == "Concluido"]),
                "overdue": len(overdue),
                "rate": completion_rate(related),
                "critical": critical,
            }
        )
    return render_template("monitoring_plans.html", plans=plans)


@app.route("/monitoring/kanban")
@login_required
def monitoring_kanban():
    actions = query_actions()
    columns = {status: [] for status in STATUS_OPTIONS}
    for action in actions:
        columns.setdefault(action["status"], []).append(action)
    return render_template("monitoring_kanban.html", columns=columns)


@app.route("/actions/new", methods=["GET", "POST"])
@write_required
def action_new():
    if request.method == "POST":
        try:
            save_action()
        except ValueError as exc:
            flash(str(exc))
            return redirect(url_for("action_new"))
        except sqlite3.IntegrityError:
            flash("Ja existe uma acao com esse codigo.")
            return redirect(url_for("action_new"))
        flash("Acao cadastrada com sucesso.")
        return redirect(url_for("actions_index"))
    return render_template("action_form.html", action=None, actions=query_actions(), people=get_people(), status_options=STATUS_OPTIONS)


@app.route("/actions/<int:action_id>/edit", methods=["GET", "POST"])
@write_required
def action_edit(action_id: int):
    if request.method == "POST":
        try:
            save_action(action_id)
        except ValueError as exc:
            flash(str(exc))
            return redirect(url_for("action_edit", action_id=action_id))
        except sqlite3.IntegrityError:
            flash("Ja existe outra acao com esse codigo.")
            return redirect(url_for("action_edit", action_id=action_id))
        flash("Acao atualizada com sucesso.")
        return redirect(url_for("action_detail", action_id=action_id))
    action = db().execute("SELECT * FROM actions WHERE id = ?", (action_id,)).fetchone()
    selected = [row["person_id"] for row in db().execute("SELECT person_id FROM action_people WHERE action_id = ?", (action_id,)).fetchall()]
    return render_template(
        "action_form.html",
        action=action,
        selected_people=selected,
        actions=query_actions(),
        people=get_people(),
        status_options=STATUS_OPTIONS,
    )


def save_action(action_id: int | None = None) -> None:
    form = request.form
    code = normalize_code(form.get("code"))
    if action_level(code) > 3 or code.count(".") + 1 > 3:
        raise ValueError("A hierarquia permite no maximo tres niveis.")
    parent_id = form.get("parent_id") or None
    level = action_level(code)
    if parent_id:
        parent = db().execute("SELECT code, level FROM actions WHERE id = ?", (parent_id,)).fetchone()
        if parent:
            if parent["level"] >= 3:
                raise ValueError("Uma acao de terceiro nivel nao pode receber subatividades.")
            level = min(action_level(parent["code"]) + 1, 3)
    fields = (
        code,
        parent_id,
        level,
        form.get("title"),
        form.get("area"),
        form.get("why"),
        form.get("planned_start") or None,
        form.get("planned_end") or None,
        form.get("actual_start") or None,
        form.get("actual_end") or None,
        form.get("place"),
        form.get("method"),
        form.get("cost"),
        form.get("status") or "Nao iniciado",
        form.get("observations"),
    )
    if action_id:
        db().execute(
            """
            UPDATE actions SET code=?, parent_id=?, level=?, title=?, area=?, why=?, planned_start=?,
                planned_end=?, actual_start=?, actual_end=?, place=?, method=?, cost=?, status=?,
                observations=?, updated_at=CURRENT_TIMESTAMP
            WHERE id=?
            """,
            (*fields, action_id),
        )
    else:
        cur = db().execute(
            """
            INSERT INTO actions (
                code, parent_id, level, title, area, why, planned_start, planned_end,
                actual_start, actual_end, place, method, cost, status, observations
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            fields,
        )
        action_id = cur.lastrowid

    db().execute("DELETE FROM action_people WHERE action_id = ?", (action_id,))
    for person_id in form.getlist("people"):
        db().execute("INSERT OR IGNORE INTO action_people (action_id, person_id) VALUES (?, ?)", (action_id, person_id))
    db().commit()


@app.route("/actions/<int:action_id>")
@login_required
def action_detail(action_id: int):
    action = db().execute(
        """
        SELECT a.*, p.title AS parent_title, GROUP_CONCAT(pe.name, ' | ') AS people
        FROM actions a
        LEFT JOIN actions p ON p.id = a.parent_id
        LEFT JOIN action_people ap ON ap.action_id = a.id
        LEFT JOIN people pe ON pe.id = ap.person_id
        WHERE a.id = ?
        GROUP BY a.id
        """,
        (action_id,),
    ).fetchone()
    updates = db().execute("SELECT * FROM updates WHERE action_id = ? ORDER BY update_date DESC, id DESC", (action_id,)).fetchall()
    children = query_actions({"q": ""})
    children = [c for c in children if c["parent_id"] == action_id]
    return render_template("action_detail.html", action=action, updates=updates, children=children, status_options=STATUS_OPTIONS)


@app.route("/actions/<int:action_id>/updates", methods=["POST"])
@write_required
def update_new(action_id: int):
    form = request.form
    status = form.get("status") or "Em andamento"
    db().execute(
        """
        INSERT INTO updates (action_id, update_date, status, progress, comment, next_step, blocker)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            action_id,
            form.get("update_date") or date.today().isoformat(),
            status,
            int(form.get("progress") or 0),
            form.get("comment"),
            form.get("next_step"),
            form.get("blocker"),
        ),
    )
    db().execute("UPDATE actions SET status = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?", (status, action_id))
    db().commit()
    flash("Acompanhamento registrado.")
    return redirect(url_for("action_detail", action_id=action_id))


@app.route("/people", methods=["GET", "POST"])
@login_required
def people_index():
    if request.method == "POST":
        if not g.current_user or g.current_user["role"] not in WRITE_ROLES:
            flash("Seu perfil permite consulta, mas nao permite cadastrar responsaveis.")
            return redirect(url_for("people_index"))
        name = request.form.get("name", "").strip()
        if name:
            db().execute(
                "INSERT OR IGNORE INTO people (name, email, area) VALUES (?, ?, ?)",
                (name, request.form.get("email"), request.form.get("area")),
            )
            db().commit()
            flash("Responsavel cadastrado.")
        return redirect(url_for("people_index"))
    rows = db().execute(
        """
        SELECT p.*, COUNT(ap.action_id) AS action_count
        FROM people p
        LEFT JOIN action_people ap ON ap.person_id = p.id
        WHERE p.active = 1
        GROUP BY p.id
        ORDER BY p.name
        """
    ).fetchall()
    return render_template("people.html", people=rows)


@app.route("/import", methods=["GET", "POST"])
@admin_required
def import_page():
    if request.method == "POST":
        path = Path(request.form.get("path") or DEFAULT_IMPORT)
        count, links = import_excel(path)
        flash(f"Importacao concluida: {count} acoes e {links} vinculos de responsaveis.")
        return redirect(url_for("dashboard"))
    return render_template("import.html", default_path=DEFAULT_IMPORT)


@app.route("/export/actions.csv")
@login_required
def export_actions():
    actions = query_actions()
    output = []
    headers = ["codigo", "nivel", "acao", "responsaveis", "status", "inicio_planejado", "fim_planejado", "inicio_real", "fim_real", "observacoes"]
    output.append(headers)
    for a in actions:
        output.append([a["code"], a["level"], a["title"], "; ".join(a["people"]), a["status"], a["planned_start"], a["planned_end"], a["actual_start"], a["actual_end"], a["observations"]])

    def generate():
        from io import StringIO

        buffer = StringIO()
        writer = csv.writer(buffer, delimiter=";")
        for row in output:
            buffer.seek(0)
            buffer.truncate(0)
            writer.writerow(row)
            yield buffer.getvalue()

    return Response(generate(), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=acoes.csv"})


def cli() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--import", dest="import_path")
    args = parser.parse_args()
    init_db()
    if args.import_path:
        count, links = import_excel(Path(args.import_path))
        print(f"Importacao concluida: {count} acoes e {links} vinculos de responsaveis.")
        return
    app.run(debug=True)


if __name__ == "__main__":
    cli()
